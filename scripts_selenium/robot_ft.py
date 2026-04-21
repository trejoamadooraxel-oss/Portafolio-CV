








import common
import pandas as pd
import csv
from io import BytesIO
import openpyxl
from datetime import datetime
from zipfile import ZipFile
import os
import xlrd
import json
from types import SimpleNamespace
from pathlib import Path

import argparse
import gzip
import re
import sys
import shutil
import time
from enum import Enum
import subprocess
import requests
import base64
import codecs

#region Global

header = ''
trace = None
params = None
paths = None
fileHelper = None


class HeaderException(Exception):
    def __init__(self, code, message):
        self.code = code
        self.message = message
        super().__init__(self.message)

#endregion Global

# region Validations


def isDate(value, format):
    try:
        datetime.strptime(str(value), format)
        return True
    except Exception as e:
        return False


def isFloat(value, decimal_separator):
    return value.replace('-', '', 1).replace(decimal_separator, '', 1).isdigit()


def isInt(value):
    return value.replace('-', '', 1).isdigit()


def is_header(conf_subtype, line, errors):
    global trace
    result = True
    pos = 0
    # Trace.write("I", "check header")
    if conf_subtype.check_header == 'Y':
        # Trace.write("I", "check header 2")
        for field in conf_subtype.fields:
            # Trace.write("I", "check header 3")
            in_line = True
            name = field.file_name
            # Trace.write("I", "check header 4")
            pos = field.pos
            # Trace.write("I", f"Field: {name}")
            if pos >= len(line):
                if pos != 999:
                    in_line = False
                    errors.append({"name": "error header",
                                   "description": f"El pos {pos} configurado supera la longitud de campos de la linea: {len(line)}",
                                   "index": 0})
            else:
                in_line = name == line[pos] or name == ''
                if not in_line:
                    errors.append({"name": "error header",
                                   "description": f"El campo {name} no corresponde con el header: {line[pos]}",
                                   "index": 0})
            # Trace.write("I", line)
            # Trace.write("I", f"in line: {in_line}")
            result = result and in_line
    trace.write("I", f"saliendo validate header. Result: {result}, Errors: {errors}")
    return result


def validate_time_id(robot_days, conf_subtype, time_id, index, errors):
    try:
        if conf_subtype.check_date == "day" and not time_id in robot_days:
            errors.append(
                {"name": "error date", "description": f"La fecha {time_id} no corresponde con la esperada: {robot_days}",
                 "index": index})
        if conf_subtype.check_date == "mon" and robot_days[0][0:7] != time_id[0:7]:
            errors.append({"name": "error date",
                           "description": f"La fecha {time_id} no corresponde con el mes esperado: {robot_days[0][0:7]}",
                           "index": index})
    except Exception as e:
        raise Exception(f"Error validando el time_id: {time_id}. Exception: {repr(e)}")


def get_validate_field(ws, conf_subtype, line, field, index, decimal_separator, errors):
    item_val = ''
    try:
        if field.pos == 999:
            match conf_subtype.format_type:
                case 'xls':
                    item_val = ws.iloc[field.cell_row, field.cell_column]
                case 'xlsx':
                    item_val = ws.cell(row=field.cell_row, column=field.cell_column).value
                case _:
                    errors.append({"name": "error pos 999",
                                   "description": f"El pos de un field solo puede estar configurado con 999 para archivos xls o xlsx",
                                   "index": index})
        else:
            item_val = line[field.pos]
        if hasattr(field, 'transforms'):
            for transform in field.transforms:
                match transform.type:
                    case 'split':
                        item_val = item_val.split(transform.character)[transform.index]
                    case 'strip':
                        item_val = item_val.strip()
                    case 'replace':
                        item_val = item_val.replace(transform.prev, transform.new)
                    case _:
                        pass
        if hasattr(field, 'transform'):
            match field.transform.type:
                case 'split':
                    item_val = item_val.split(field.transform.character)[field.transform.index]
                case 'strip':
                    item_val = item_val.strip()
                case 'replace':
                    item_val = item_val.replace(field.transform.prev, field.transform.new)
                case _:
                    pass
        if (field.required == 'Y' and (item_val == '' or item_val == None)):
            errors.append(
                {"name": "error required", "description": f"El campo {field.bd_name} es requerido!", "index": index})
        if (field.bd_type == 'FLOAT64'):
            if field.file_type in ['int', 'float']:
                if not isinstance(item_val, (int, float)) and item_val != None:
                    errors.append({"name": "error type",
                                   "description": f"El campo {field.bd_name} no es del tipo: {field.bd_type}, value: {item_val}",
                                   "index": index})
            else:
                if not isFloat(str(item_val), decimal_separator) and item_val != None and item_val != '':
                    errors.append({"name": "error type",
                                   "description": f"El campo {field.bd_name} no es del tipo: {field.bd_type}, value: {item_val}",
                                   "index": index})
        if (field.bd_type == 'DATE'):
            if field.file_type == 'date':
                if not isinstance(item_val, datetime) and item_val != None:
                    errors.append({"name": "error type",
                                   "description": f"El campo {field.bd_name} no es del tipo: {field.bd_type} con formato:{field.file_format}, value: {item_val}",
                                   "index": index})
                else:
                    item_val = item_val.strftime("%Y-%m-%d")
            else:
                if not isDate(item_val, field.file_format) and item_val != None and item_val != '':
                    errors.append({"name": "error type",
                                   "description": f"El campo {field.bd_name} no es del tipo: {field.bd_type} con formato:{field.file_format}, value: {item_val}",
                                   "index": index})
                else:
                    if item_val != '':
                        item_val = datetime.strptime(str(item_val), field.file_format).strftime("%Y-%m-%d")
    except Exception as e:
        raise Exception(f"Line: {line}, Field: {field}, index: {index}. Exception: {repr(e)}")
    return item_val

# endregion Validations

# region File Types


def process_xls(conf_subtype, files, rows_src, rows_gen, robot_days, data_rows, preview, errors, days, zip_type):
    trace.write("I", "Procesando XLS")
    filename_gen = f"{os.path.splitext(files[0])[0]}_parsed.csv"
    trace.write("I", f"Filename gen: {filename_gen}")
    with open(f"{paths.transformed}{filename_gen}", 'a', encoding='utf-8', newline='') as f_new_file:
        csv_writer = csv.writer(f_new_file)
        for file in files:
            match zip_type:
                case common.ZipType.ZIP:
                    input_zip = ZipFile(paths.source + file)
                    input_zip.extractall(paths.source)

                    file_to_transform = [f for f in os.listdir(paths.source) if
                            f.endswith('.csv') or f.endswith('.CSV') or f.endswith('.Csv')][0]
                case common.ZipType.NONE:
                    file_to_transform = file
                case _:
                    file_to_transform = file

            trace.write("I", "archivo bytes....")
            header = ''
            workbook = xlrd.open_workbook_xls(f"{paths.source}{file_to_transform}", ignore_workbook_corruption=True)
            #df = pd.read_excel(f"{paths.source}{file_to_transform}", sheet_name=None, header=None, dtype=str)
            df = pd.read_excel(workbook, sheet_name=None, header=None, dtype=str)

            sheets_xls = list(df.keys())
            trace.write("I", sheets_xls)
            trace.write("I", "archivo cargado....")
            has_header = conf_subtype.header == 'Y'
            header_offset = conf_subtype.xls_conf.header_offset
            fields = list(conf_subtype.fields)
            defaults = conf_subtype.defaults
            if hasattr(conf_subtype, 'pos_import_pivot'):
                trace.write("I", "has pivot pos")
                pivot = conf_subtype.pos_import_pivot
            else:
                pivot = ''
            trace.write("I", "ordenando campos....")
            fields.sort(key=lambda x: x.pos)
            trace.write("I", "campos ordenados....")
            sheets = []
            sheets_names_file_conf = conf_subtype.xls_conf.sheets_names
            # toma los nombres de hojas a procesar segun la configuracion #all# para todas las hojas
            if len(sheets_names_file_conf) > 0:
                if "#all#" in conf_subtype.xls_conf.sheets_names:
                    sheets = sheets_xls
                else:
                    sheets = list(set(sheets_xls) & set(conf_subtype.xls_conf.sheets_names))
                    if len(sheets) == 0:
                        errors.append({"name": "error nombre hoja",
                                       "description": f"Hojas solicitadas: {conf_subtype.xls_conf.sheets_names}, hojas en el archivo: {sheets_xls}",
                                       "index": 0})
                        return

            else:
                for index in conf_subtype.xls_conf.sheets_nums:
                    sheets.append(sheets_xls[index])
            # recorre las hojas a procesar.
            for sheet, hoja in df.items():
                if sheet in sheets:
                    trace.write("I", f"procecing sheet: {sheet}")
                    ws = hoja
                    columns = list(ws.columns)
                    index_header = 0
                    i = 0
                    pos_import = ''
                    # Trace.write("I",f"columnas: {columns}")
                    max_row = len(ws)
                    trace.write("I", "Max rows:  " + str(max_row))
                    for row in ws.itertuples(index=False):
                        line = list(row)
                        #trace.write("I", f"line: {line}")
                        i += 1
                        if i > (max_row - conf_subtype.xls_conf.ignore_last_rows):
                            pass
                        else:
                            data_rows = []
                            if not hasattr(conf_subtype.xls_conf, 'ignore_none_col_id') or (
                                    hasattr(conf_subtype.xls_conf, 'ignore_none_col_id') and str(
                                    line[conf_subtype.xls_conf.ignore_none_col_id]) != 'nan'):
                                # Trace.write("I", f"procesing line: {i}")
                                # process_line(line, i, rows_src, time_id, filepath, fields, defaults, has_header, header_offset, data_rows, preview, errors, sheet, pivot)
                                process_line(ws, conf_subtype, line, i, file, rows_src, rows_gen, robot_days, fields, defaults,
                                             has_header, header_offset, data_rows, preview, errors, days, sheet, pivot)
                                if (len(errors) == 0):
                                    csv_writer.writerows(data_rows)


def process_xlsx(conf_subtype, files, rows_src, rows_gen, robot_days, data_rows, preview, errors, days, zip_type):
    trace.write("I", "Procesando XLSX")
    ignore_last_rows = conf_subtype.xls_conf.ignore_last_rows
    filename_gen = f"{os.path.splitext(files[0])[0]}_parsed.csv"
    trace.write("I", f"Filename gen: {filename_gen}")

    if hasattr(params, 'pos_import_pivot'):
        trace.write("I", "has pivot pos")
        pivot = conf_subtype.pos_import_pivot
    else:
        pivot = ''

    with open(f"{paths.transformed}{filename_gen}", 'a', encoding='utf-8', newline='') as f_new_file:
        csv_writer = csv.writer(f_new_file)
        for file in files:
            match zip_type:
                case common.ZipType.ZIP:
                    input_zip = ZipFile(paths.source + file)
                    input_zip.extractall(paths.source)

                    file_to_transform = [f for f in os.listdir(paths.source) if
                            f.endswith('.csv') or f.endswith('.CSV') or f.endswith('.Csv')][0]
                case common.ZipType.NONE:
                    file_to_transform = file
                case _:
                    file_to_transform = file
            #blob = common.FileHelper.get_file_from_gs(file, bucket)
            #file_blob = BytesIO(blob.download_as_string())
            trace.write("I", "archivo bytes....")
            wb = openpyxl.load_workbook(f"{paths.source}{file_to_transform}")
            trace.write("I", "archivo cargado....")
            print(wb.sheetnames)
            has_header = conf_subtype.header == 'Y'
            header_offset = conf_subtype.xls_conf.header_offset
            fields = conf_subtype.fields
            defaults = conf_subtype.defaults
            fields.sort(key=lambda x: x.pos)
            sheets = []
            sheets_names_file_conf = conf_subtype.xls_conf.sheets_names
            # toma los nombres de hojas a procesar segun la configuracion #all# para todas las hojas
            if len(sheets_names_file_conf) > 0:
                if "#all#" in conf_subtype.xls_conf.sheets_names:
                    sheets = wb.sheetnames
                else:
                    sheets = list(set(wb.sheetnames) & set(conf_subtype.xls_conf.sheets_names))
            else:
                for index in conf_subtype.xls_conf.sheets_nums:
                    sheets.append(wb.sheetnames[index])
            # recorre las hojas a procesar.
            for sheet in sheets:
                trace.write("I", f"procecing sheet: {sheet}")
                ws = wb[sheet]
                # pos_import = ws.cell(row=3, column=1).value
                # Trace.write("I", f"pos_import: {pos_import}")

                index_header = 0
                i = 0
                trace.write("I", "Max rows:  " + str(ws.max_row))
                max_row = ws.max_row
                for line in ws.iter_rows(values_only=True):
                    # print(line)
                    i += 1
                    if i > (max_row - ignore_last_rows):
                        pass
                    else:
                        data_rows = []
                        if not hasattr(conf_subtype.xls_conf, 'ignore_none_col_id') or (
                                hasattr(conf_subtype.xls_conf, 'ignore_none_col_id') and line[conf_subtype.xls_conf.ignore_none_col_id] != None):
                            # Trace.write("I", f"procesing line: {i}")
                            process_line(ws, conf_subtype, line, i, file, rows_src, rows_gen, robot_days,
                                         fields, defaults, has_header, header_offset, data_rows, preview,
                                         errors, days, sheet, pivot)
                            if (len(errors) == 0):
                                csv_writer.writerows(data_rows)


def process_csv(conf_subtype, files, rows_src, rows_gen, robot_days, data_rows, preview, errors, days, zip_type):
    trace.write("I", "Procesando CSV")
    bucket = params.bucket_src

    has_header = conf_subtype.header == 'Y'
    fields = conf_subtype.fields
    # print(fields)
    defaults = conf_subtype.defaults
    fields.sort(key=lambda x: x.pos)
    if hasattr(conf_subtype.csv_conf, 'ignore_len_row'):
        ignore_len_row = conf_subtype.csv_conf.ignore_len_row
    else:
        ignore_len_row = -1

    index_header = 0
    i = 0
    filename_gen = f"{os.path.splitext(files[0])[0]}_parsed.csv"
    trace.write("I", f"creando archivo {filename_gen}")

    with open(f"{paths.transformed}{filename_gen}", 'a', encoding='utf-8', newline='') as f_new_file:
        csv_writer = csv.writer(f_new_file)
        for file in files:
            match zip_type:
                case common.ZipType.ZIP:
                    input_zip = ZipFile(paths.source + file)
                    input_zip.extractall(paths.source)

                    file_to_transform = [f for f in os.listdir(paths.source) if
                            f.endswith('.csv') or f.endswith('.CSV') or f.endswith('.Csv')][0]
                case common.ZipType.NONE:
                    file_to_transform = file
                case _:
                    file_to_transform = file

            with open(f"{paths.source}{file_to_transform}", "r", encoding=conf_subtype.csv_conf.encoding) as f_input:
                csv_reader = csv.reader(fix_nulls(f_input), delimiter=conf_subtype.csv_conf.delimiter)
                for j, line in enumerate(csv_reader):
                    # print(line)
                    i += 1
                    data_rows = []
                    if ignore_len_row == -1 or len(line) > ignore_len_row:
                        process_line('', conf_subtype, line, i, file, rows_src, rows_gen,
                                     robot_days, fields, defaults, has_header, 0, data_rows, preview,
                                     errors, days, '', '')
                    if (len(errors) == 0):
                        # Trace.write("I", data_rows)
                        csv_writer.writerows(data_rows)

# endregion File Types

# region Functions


def fix_nulls(s):
    for line in s:
        yield line.replace('\0', '')


def get_time_id(ws, conf_subtype, robot_days, line, fields, index, errors):
    time_id = ""
    time_day = ""
    time_month = ""
    time_year = ""
    for i in range(len(fields)):
        field = fields[i]
        if hasattr(field, "role"):
            pos_field = field.pos
            role = field.role
            if role == "time_id":
                time_id = get_validate_field(ws, conf_subtype, line, field, index, '', errors)
            if role == "time_day":
                time_day = str(line[pos_field]).rjust(2, '0')
            if role == "time_month":
                time_month = str(line[pos_field]).rjust(2, '0')
            if role == "time_year":
                time_year = str(line[pos_field])
        # else:
        # Trace.write("I", "no role")

    if time_id != "":
        return time_id
    if time_day != "":
        return time_year + "-" + time_month + "-" + time_day
    return robot_days[0]


def get_data_line(ws, conf_subtype, fields, defaults, errors, days, line, index, filename, chain_id, supplier_id, robot_days, pivot):
    global header, trace, params
    row = []
    rows = []
    preview_row = {}
    result = {}
    row.append(filename)
    row.append(chain_id)
    row.append(supplier_id)
    time_id = get_time_id(ws, conf_subtype, robot_days, line, fields, index, errors)
    validate_time_id(robot_days, conf_subtype, time_id, index, errors)
    row.append(time_id)
    days.add(time_id)

    pos_field = 0

    for field in fields:
        pos = field.pos
        item = ''
        if pos != 999:
            item = line[pos]
        value_item = get_validate_field(ws, params, line, field, index, conf_subtype.decimal_separator, errors)
        row.append(value_item)
        preview_row[field.bd_name] = value_item

    for item in defaults:
        row.append(item.value)
        preview_row[item.bd_name] = item.value
    if pivot != '':
        start_pos = pivot.start_pos
        end_pos = len(line) - pivot.end_offset
        for pos_i in range(start_pos, end_pos):
            new_row = row.copy()
            trace.write("I", f"Pos: {pos_i}, pos_import: {header[pos_i]}, value: {line[pos_i]}")
            new_row.append(header[pos_i])
            new_row.append(line[pos_i])
            preview_row[params.pos_import_pivot.header_bd_name] = header[pos_i]
            preview_row[params.pos_import_pivot.item_bd_name] = line[pos_i]
            rows.append(new_row)
    else:
        rows.append(row)

    result["rows"] = rows
    result["preview_row"] = preview_row
    return result


def process_line(ws, conf_subtype, line, i, filename, rows_src, rows_gen, robot_days, fields, defaults, has_header,
                 header_offset, data_rows, preview, errors, days, sheet, pivot):
    global header, trace, params
    rows_src += 1
    if i == (1 + header_offset) and has_header:
        if is_header(conf_subtype, line, errors):
            header = line
            if len(line) < conf_subtype.header_cols_num:
                if sheet != '':
                    errors.append({"name": "error header",
                                   "description": f"Error en el header. Hay menos columnas de las esperadas.Sheet: {sheet}",
                                   "index": 0})
                else:
                    errors.append({"name": "error header",
                                   "description": "Error en el header. Hay menos columnas de las esperadas!",
                                   "index": 0})
                raise HeaderException(common.ResultType.HEADER_ERROR.value, 'Error en el header')
        if len(errors) > 0:
            raise HeaderException(common.ResultType.HEADER_ERROR.value, 'Error en el header')
    if (has_header and i > (1 + header_offset)) or not has_header:
        try:
            data = get_data_line(ws, conf_subtype, fields, defaults, errors, days, line, i, filename,
                                 params.chain_id, params.supplier_id, robot_days,
                                 pivot)
            data_rows.extend(data['rows'])
            rows_gen += 1
            if len(preview) < 100:
                preview.append(data['preview_row'])
        except Exception as e:
            trace.write("E", f"Excepcion. Line: {i}. Error: {e}")
            errors.append({"name": "error no controlado", "description": f"Error: {e}", "index": 0})

# endregion Functions

# region Main


def transform_report(_trace, _params, _paths, _fileHelper, current_subtask, zip_type):
    global trace, params, paths, fileHelper
    trace = _trace
    params = _params
    paths = _paths
    fileHelper = _fileHelper
    try:
        fileHelper.clear_paths([paths.transformed])
        subtype = current_subtask.subtype
        trace.write('I', f'Iniciando proceso de transformación. Subtype: {subtype}')
        conf_subtype = list(filter(lambda x: x.subtype == subtype, params.subtypes_conf))[0]
        rows_src = 0
        rows_gen = 0
        files = current_subtask.files_src
        robot_days = current_subtask.days
        i = 0

        errors = []
        preview = []
        data_rows = []
        days = set()

        match conf_subtype.format_type:
            case 'xls':
                process_xls(conf_subtype, files, rows_src, rows_gen, robot_days, data_rows, preview, errors, days, zip_type)
            case 'xlsx':
                process_xlsx(conf_subtype, files, rows_src, rows_gen, robot_days, data_rows, preview, errors, days, zip_type)
            case 'csv':
                process_csv(conf_subtype, files, rows_src, rows_gen, robot_days, data_rows, preview, errors, days, zip_type)
        filename_gen = f"{os.path.splitext(files[0])[0]}_parsed.csv"

        if len(errors) > 0:
            current_subtask.set_gen("", "", errors, 0, 0)
            current_subtask.set_status("E", common.ResultType.TRANSFORM_CHECK, "Hay errores en la validacion de la transformacion" )

        else:
            #send_file_name = fileHelper.send_to_gs(paths.transformed, filename_gen.replace(".csv", ""), 'csv', params.bucket_gen, True)
            #send_file_ext = os.path.splitext(send_file_name)[1]
            trace.write('I', 'Proceso de Transformacion Finalizado.')
            # los diferentes dias que tiene como time_id. Generalmente son el mismo de la subtask
            if len(days) > 0:
                current_subtask.set_days(days)
            #current_subtask.set_gen(send_file_name, send_file_ext, errors, rows_src, rows_gen)

    except Exception as e:
        trace.write('E', f'Error: Finalizado en Area Transformacion. Excepcion: {repr(e)}')
        raise common.TransformException(common.ResultType.TRANSFORM_ERROR.value, repr(e))

# endregion Main